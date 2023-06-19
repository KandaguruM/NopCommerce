import openpyxl


file = "C:/Users/ADMIN/Downloads/Nop Commerce Testcase.xlsx"
sheetname = "REGISTER"
column = 11


def max_row(file, sheetname):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetname]
        return sheet.max_row


def max_column(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_row


def Write_data(file, sheetname, row, column, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row, column).value = data
    workbook.save(file)
